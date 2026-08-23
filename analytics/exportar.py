"""Exportacao generica de tabelas de analise para Excel.

Cada pagina de analise (ABC, Oportunidades, ...) tem colunas e formulas
diferentes — este modulo so cuida do que e comum: layout do cabecalho,
congelamento de painel, e a aba "Como foi calculado" com as premissas junto,
porque a planilha circula solta por e-mail, longe da tela que explica os
numeros. O mesmo padrao usado em compra.py::exportar_xlsx, generalizado.
"""
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_TITULO = Font(bold=True, color="FFFFFF")
_FUNDO = PatternFill("solid", fgColor="7A1420")
_NEGRITO = Font(bold=True)


@dataclass
class ColunaXlsx:
    cabecalho: str
    chave: str
    largura: int = 18
    formato: str | None = None
    valor: Callable[[dict], Any] | None = None

    def ler(self, linha: dict) -> Any:
        return self.valor(linha) if self.valor else linha.get(self.chave)


@dataclass
class AbaXlsx:
    nome: str
    colunas: list[ColunaXlsx]
    linhas: list[dict]


@dataclass
class SecaoCalculo:
    titulo: str
    formula: str = ""
    valores: dict[str, Any] = field(default_factory=dict)
    premissas: list[str] = field(default_factory=list)


def _aba_tabela(wb: Workbook, aba: AbaXlsx) -> None:
    ws = wb.create_sheet(aba.nome[:31])  # 31 = limite do Excel para nome de aba
    for col, c in enumerate(aba.colunas, start=1):
        cel = ws.cell(row=1, column=col, value=c.cabecalho)
        cel.font, cel.fill = _TITULO, _FUNDO
        cel.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = c.largura
    ws.freeze_panes = "A2"
    for r, linha in enumerate(aba.linhas, start=2):
        for col, c in enumerate(aba.colunas, start=1):
            cel = ws.cell(row=r, column=col, value=c.ler(linha))
            if c.formato:
                cel.number_format = c.formato
    if not aba.linhas:
        ws.cell(row=2, column=1, value="Sem dados neste recorte.")


def _aba_premissas(wb: Workbook, titulo_geral: str, secoes: list[SecaoCalculo]) -> None:
    wp = wb.create_sheet("Como foi calculado")
    wp.column_dimensions["A"].width = 34
    wp.column_dimensions["B"].width = 96
    linha = 1

    def par(a: str, b: Any, negrito: bool = False) -> None:
        nonlocal linha
        ca = wp.cell(row=linha, column=1, value=a)
        cb = wp.cell(row=linha, column=2, value=str(b) if b is not None else "")
        cb.alignment = Alignment(wrap_text=True, vertical="top")
        if negrito:
            ca.font = cb.font = _NEGRITO
        linha += 1

    par(titulo_geral, "", True)
    linha += 1
    for s in secoes:
        par(s.titulo, "", True)
        if s.formula:
            par("Fórmula", s.formula)
        for rot, val in s.valores.items():
            par(rot, val)
        for p in s.premissas:
            par("", p)
        linha += 1


def montar_workbook(titulo_geral: str, abas: list[AbaXlsx],
                    secoes_calculo: list[SecaoCalculo]) -> bytes:
    """Uma aba por tabela + uma aba final 'Como foi calculado' com as
    premissas de todas as secoes juntas."""
    if not abas:
        raise ValueError("Nada para exportar neste recorte.")
    wb = Workbook()
    wb.remove(wb.active)
    for aba in abas:
        _aba_tabela(wb, aba)
    _aba_premissas(wb, titulo_geral, secoes_calculo)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def nome_arquivo_seguro(texto: str) -> str:
    return "".join(c if c.isalnum() or c in "-_" else "_" for c in texto)
