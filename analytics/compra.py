"""Sugestao de pedido de compra para o distribuidor.

A conta central e a do comprador, nao uma projecao estatistica:

    necessidade (un) = venda diaria x DDE alvo - estoque disponivel - pendencia

Onde:
  venda diaria    vem do sell-out do periodo selecionado OU da media do
                  proprio arquivo de estoque (o usuario escolhe, mesmo
                  parametro base_velocidade do resto da tela de estoque)
  DDE alvo        e do usuario: por agrupamento, com override por SKU
  pendencia       e o que ja esta comprado e nao chegou. Ignorar isso faz
                  o sistema mandar comprar duas vezes a mesma coisa.

SOBRE "CATEGORIA": a base importada nao tem campo de categoria terapeutica.
molecula e fabricante vieram vazios no cadastro; marca existe, mas mistura
marca de verdade (FORDAY, DORALEX) com nome de molecula (PARACETAMOL MG).
Entao esta tela oferece os agrupamentos que o dado SUSTENTA — curva ABC,
faixa de cobertura e marca — e diz qual esta em uso. Inventar categorias
seria opiniao disfarcada de dado.

Quando o usuario define um VALOR ALVO de compra, o pedido nao e inflado nem
cortado no meio: os SKUs entram por ordem de urgencia (menor cobertura
primeiro) ate o orcamento acabar, com teto por SKU para o pedido nao virar
um caminhao de um item so. Se a necessidade real for menor que o alvo, o
sistema diz isso em vez de completar com o que nao precisa.
"""
import math
import sqlite3
from typing import Any

from . import abc as abc_mod
from . import estoque as estoque_mod
from .contexto import carregar
from .formulas import Calculo

# Agrupamentos possiveis. Todos derivados do dado, nenhum inventado.
AGRUPAMENTOS = ("abc", "estoque", "marca")

DDE_ALVO_PADRAO = 60.0

# Teto por SKU como fracao do orcamento, quando ha valor alvo. Sem teto, um
# unico SKU de giro alto consome o pedido inteiro — foi o teto de 20% que o
# proprio dossie da EMEFARMA usou.
TETO_POR_SKU_PADRAO = 0.20

# Ordem de gravidade da faixa de cobertura, para o desempate de urgencia.
_ORDEM_CLASSE = ["SAUDAVEL", "ATENCAO", "ALTO", "CRITICO", "ZUMBI", "INDEFINIDO"]

# Na base real, 51 dos 223 SKUs de estoque nao existem no sell-out sob o mesmo
# produto_id: as duas fontes trazem EANs diferentes para o mesmo item fisico
# (DIPIRONA 1G aparece como 7898700413007 no estoque e 7898049797615 no
# sell-out). Esses SKUs nao tem curva ABC nem marca, e precisam de DDE manual.
_GRUPO_SEM_LIGACAO = "Sem ligação com o sell-out"


def _sem(motivo: str) -> dict:
    return {"disponivel": False, "motivo": motivo}


def _pendencias(con: sqlite3.Connection, client_id: int, data_ref: str,
                filial: str | None) -> dict[int, float]:
    where = ["client_id = ?", "data_ref = ?"]
    params: list[Any] = [client_id, data_ref]
    if filial is not None:
        where.append("filial = ?")
        params.append(filial)
    return {pid: float(p or 0) for pid, p in con.execute(
        f"SELECT produto_id, sum(coalesce(pendencia_un,0)) FROM v_estoque"
        f" WHERE {' AND '.join(where)} GROUP BY produto_id", params)}


def sugerir_pedido(con: sqlite3.Connection, client_id: int, ini: int, fim: int, *,
                   agrupamento: str = "abc",
                   dde_por_grupo: dict[str, float] | None = None,
                   dde_padrao: float = DDE_ALVO_PADRAO,
                   dde_por_produto: dict[int, float] | None = None,
                   base_velocidade: str = "fonte",
                   filial: str | None = None,
                   valor_alvo: float | None = None,
                   teto_por_sku: float = TETO_POR_SKU_PADRAO,
                   incluir_sem_giro: bool = False) -> dict:
    """Monta a sugestao de pedido.

    dde_por_grupo   {nome do grupo -> dias}. O que nao vier usa dde_padrao.
    dde_por_produto {produto_id -> dias}. Vence sobre o grupo — e o ajuste
                    linha a linha que o comprador faz na hora.
    valor_alvo      teto de compra em R$. None = pedido pela necessidade cheia.
    incluir_sem_giro  SKU sem venda no periodo tem DDE indefinido; por padrao
                    fica fora (comprar mais do que nao gira e o oposto do
                    objetivo), mas aparece na contagem.
    """
    if agrupamento not in AGRUPAMENTOS:
        raise ValueError(
            f"agrupamento invalido: {agrupamento}. Use um de: {', '.join(AGRUPAMENTOS)}.")
    if dde_padrao <= 0:
        raise ValueError("O DDE alvo precisa ser maior que zero.")
    if valor_alvo is not None and valor_alvo <= 0:
        raise ValueError("O valor alvo precisa ser maior que zero.")
    if not 0 < teto_por_sku <= 1:
        raise ValueError("O teto por SKU e uma fracao entre 0 e 1.")

    disp = carregar(con, client_id)
    if not disp.tem_sellout:
        return _sem(disp.motivo_indisponivel)

    pos = estoque_mod.posicao(con, client_id, ini, fim, base_velocidade=base_velocidade,
                              filial=filial, limite=1000000)
    if not pos.get("disponivel"):
        return pos

    dde_por_grupo = dde_por_grupo or {}
    dde_por_produto = dde_por_produto or {}

    # Grupo de cada produto, no criterio escolhido.
    grupo_de: dict[int, str] = {}
    if agrupamento == "abc":
        curva = abc_mod.curva_abc(con, client_id, ini, fim, com_cobertura=False)
        if curva.get("disponivel"):
            for i in curva["itens"]:
                grupo_de[i["produto_id"]] = f"Curva {i['classe_abc']}"
    elif agrupamento == "marca":
        for pid, marca in con.execute(
            "SELECT id, marca FROM dim_product WHERE marca IS NOT NULL"):
            grupo_de[pid] = marca

    pend = _pendencias(con, client_id, pos["data_ref"], filial)

    linhas: list[dict] = []
    n_sem_giro = 0
    # Um produto pode ter posicao em mais de uma filial; o pedido e por SKU.
    por_produto: dict[int, dict] = {}
    for i in pos["itens"]:
        d = por_produto.setdefault(i["produto_id"], {
            "produto_id": i["produto_id"], "produto": i["produto"],
            "estoque_un": 0.0, "valor_estoque": 0.0, "venda_dia": 0.0,
            "custo": None, "classe_estoque": None, "filiais": [],
        })
        d["estoque_un"] += i["estoque_disp_un"]
        d["valor_estoque"] += i["valor_estoque"]
        vd = (i["venda_dia_fonte"] if base_velocidade == "fonte"
              else i["venda_dia_periodo"]) or 0.0
        d["venda_dia"] += vd
        if d["custo"] is None and i["custo_reposicao"]:
            d["custo"] = i["custo_reposicao"]
        c = i["classificacao"]
        if d["classe_estoque"] is None or (
                c in _ORDEM_CLASSE and d["classe_estoque"] in _ORDEM_CLASSE
                and _ORDEM_CLASSE.index(c) > _ORDEM_CLASSE.index(d["classe_estoque"])):
            d["classe_estoque"] = c
        d["filiais"].append(i["filial"])

    for pid, d in por_produto.items():
        if d["venda_dia"] <= 0:
            n_sem_giro += 1
            if not incluir_sem_giro:
                continue
        if agrupamento == "estoque":
            grupo = d["classe_estoque"] or "INDEFINIDO"
        else:
            # SKU que existe no arquivo de estoque mas nao no sell-out (EANs
            # divergentes entre as duas fontes) nao tem curva ABC nem marca no
            # cadastro. Vai para um grupo NOMEADO, nao para um balde anonimo:
            # sao justamente os que precisam de decisao manual de DDE.
            grupo = grupo_de.get(pid) or _GRUPO_SEM_LIGACAO
        alvo = dde_por_produto.get(pid, dde_por_grupo.get(grupo, dde_padrao))
        estoque_alvo = d["venda_dia"] * alvo
        pendente = pend.get(pid, 0.0)
        # A pendencia ja esta comprada: entra como estoque futuro, senao o
        # sistema manda comprar de novo o que ja esta a caminho.
        # Para ATINGIR o DDE alvo e preciso pelo menos essa quantidade — daí
        # teto, nao arredondamento: 7.900,7 unidades viraria 7.900 e o pedido
        # entregaria menos dias do que o comprador pediu. E ninguem compra
        # fracao de unidade.
        necessidade = math.ceil(estoque_alvo - d["estoque_un"] - pendente)
        if necessidade <= 0:
            continue
        custo = d["custo"] or (d["valor_estoque"] / d["estoque_un"]
                               if d["estoque_un"] else 0.0)
        dde_atual = d["estoque_un"] / d["venda_dia"] if d["venda_dia"] else None
        linhas.append({
            "produto_id": pid, "produto": d["produto"],
            "grupo": grupo,
            "classe_estoque": d["classe_estoque"],
            "filiais": sorted(set(f for f in d["filiais"] if f)),
            "estoque_atual_un": d["estoque_un"],
            "pendencia_un": pendente,
            "venda_dia": d["venda_dia"],
            "venda_mes": d["venda_dia"] * estoque_mod.DIAS_MES,
            "dde_atual": dde_atual,
            "dde_alvo": alvo,
            "dde_origem": ("produto" if pid in dde_por_produto else
                           "grupo" if grupo in dde_por_grupo else "padrao"),
            "estoque_alvo_un": estoque_alvo,
            "sugestao_un": necessidade,
            "custo_unitario": custo,
            "sugestao_valor": necessidade * custo,
            "dde_apos_pedido": alvo,
        })

    # Urgencia: quem tem menos dias de cobertura hoje compra primeiro. SKU sem
    # cobertura calculavel vai para o fim, nao para a frente.
    linhas.sort(key=lambda x: (x["dde_atual"] if x["dde_atual"] is not None else 1e9))

    necessidade_total = sum(x["sugestao_valor"] for x in linhas)
    corte = None
    if valor_alvo is not None:
        teto_valor = valor_alvo * teto_por_sku
        avisos_orcamento: list[str] = []

        # Fase 1 (protecao): em ordem de urgencia, cada SKU recebe o minimo
        # entre sua necessidade, o teto por SKU e o que resta do orcamento. O
        # teto evita que o SKU mais urgente sozinho consuma o pedido inteiro.
        gasto = 0.0
        for x in linhas:
            necessidade_cheia = x["sugestao_valor"]
            x["sugestao_valor_cheio"] = necessidade_cheia
            x["limitado_por_teto"] = necessidade_cheia > teto_valor
            restante = valor_alvo - gasto
            alvo_fase1 = min(necessidade_cheia, teto_valor, max(0.0, restante))
            un = math.floor(alvo_fase1 / x["custo_unitario"]) if x["custo_unitario"] else 0.0
            x["sugestao_un"] = un
            x["sugestao_valor"] = un * x["custo_unitario"]
            gasto += x["sugestao_valor"]

        # Fase 2 (preenchimento): se sobrou orcamento — porque a soma das
        # necessidades tetadas foi menor que o alvo —, o restante e
        # redistribuido nos MESMOS SKUs, em ordem de urgencia, agora SEM
        # teto, ate cada um atingir a necessidade real ou o orcamento
        # acabar. Sem esta fase, dinheiro ficaria parado mesmo havendo
        # necessidade real para absorve-lo — nao trava o volume estipulado.
        sobra = valor_alvo - gasto
        n_redistribuidos = 0
        if sobra > 0.01:
            for x in linhas:
                if sobra <= 0.01:
                    break
                falta = x["sugestao_valor_cheio"] - x["sugestao_valor"]
                if falta <= 0.01 or not x["custo_unitario"]:
                    continue
                extra_un = math.floor(min(falta, sobra) / x["custo_unitario"])
                if extra_un <= 0:
                    continue
                incremento = extra_un * x["custo_unitario"]
                x["sugestao_un"] += extra_un
                x["sugestao_valor"] += incremento
                x["excedeu_teto_na_redistribuicao"] = x["sugestao_valor"] > teto_valor + 0.01
                sobra -= incremento
                gasto += incremento
                n_redistribuidos += 1

        for x in linhas:
            x.setdefault("excedeu_teto_na_redistribuicao", False)
            if x["venda_dia"] > 0 and x["custo_unitario"]:
                x["dde_apos_pedido"] = (
                    (x["estoque_atual_un"] + x["pendencia_un"] + x["sugestao_un"])
                    / x["venda_dia"])

        linhas = [x for x in linhas if x["sugestao_valor"] > 0]
        n_limitados = sum(1 for x in linhas
                          if x.get("limitado_por_teto") and not x.get("excedeu_teto_na_redistribuicao"))

        if n_redistribuidos > 0:
            avisos_orcamento.append(
                f"{n_redistribuidos} SKU(s) precisaram de mais do que o teto de "
                f"{teto_por_sku * 100:.0f}% por SKU na primeira rodada; o "
                f"orçamento que sobrou foi redistribuído para eles em ordem de "
                f"urgência, sem ultrapassar a necessidade real de cada um.")
        sobra_final = max(0.0, valor_alvo - gasto)
        if sobra_final > valor_alvo * 0.05:
            avisos_orcamento.append(
                f"Sobraram R$ {sobra_final:,.2f} do orçamento sem SKU que "
                f"precise deles neste recorte — a necessidade real da carteira "
                f"é menor que o valor alvo definido.")

        corte = {
            "valor_alvo": valor_alvo,
            "necessidade_total": necessidade_total,
            "atendido": gasto,
            "sobra_do_orcamento": sobra_final,
            "necessidade_nao_atendida": max(0.0, necessidade_total - gasto),
            "teto_por_sku_pct": teto_por_sku * 100,
            "teto_por_sku_valor": teto_valor,
            "n_limitados_por_teto": n_limitados,
            "n_redistribuidos_acima_do_teto": n_redistribuidos,
            "avisos": avisos_orcamento,
        }

    total_valor = sum(x["sugestao_valor"] for x in linhas)
    total_un = sum(x["sugestao_un"] for x in linhas)

    por_grupo: dict[str, dict] = {}
    for x in linhas:
        g = por_grupo.setdefault(x["grupo"], {
            "grupo": x["grupo"], "skus": 0, "unidades": 0.0, "valor": 0.0,
            "dde_alvo": x["dde_alvo"]})
        g["skus"] += 1
        g["unidades"] += x["sugestao_un"]
        g["valor"] += x["sugestao_valor"]

    ROTULO_AGRUP = {"abc": "curva ABC (A/B/C)",
                    "estoque": "faixa de cobertura do estoque",
                    "marca": "marca/linha do cadastro"}
    return {
        "disponivel": True,
        "data_ref": pos["data_ref"],
        "agrupamento": agrupamento,
        "base_velocidade": base_velocidade,
        "filial": filial,
        "dde_padrao": dde_padrao,
        "dde_por_grupo": dde_por_grupo,
        "n_skus": len(linhas),
        "n_sem_giro": n_sem_giro,
        "n_sem_ligacao": sum(1 for x in linhas if x["grupo"] == _GRUPO_SEM_LIGACAO),
        "valor_sem_ligacao": sum(x["sugestao_valor"] for x in linhas
                                 if x["grupo"] == _GRUPO_SEM_LIGACAO),
        "grupo_sem_ligacao": _GRUPO_SEM_LIGACAO,
        "total_unidades": total_un,
        "total_valor": total_valor,
        "necessidade_total": necessidade_total,
        "corte": corte,
        "grupos": sorted(por_grupo.values(), key=lambda g: -g["valor"]),
        "grupos_disponiveis": sorted({x["grupo"] for x in linhas}),
        "itens": linhas,
        "calculo": Calculo(
            formula=("necessidade (un) = venda diaria x DDE alvo - estoque "
                     "disponivel - pendencia; valor = necessidade x custo de "
                     "reposicao. Com valor alvo, os SKUs entram por urgencia "
                     "(menor cobertura primeiro) ate o orcamento acabar."),
            valores={
                "agrupamento": ROTULO_AGRUP[agrupamento],
                "DDE alvo padrao": f"{dde_padrao:.0f} dias",
                "base de velocidade": base_velocidade,
                "SKUs no pedido": len(linhas),
                "SKUs sem giro (fora)": n_sem_giro,
                "total do pedido": round(total_valor, 2),
                "necessidade cheia": round(necessidade_total, 2),
                "foto de estoque": pos["data_ref"],
            },
            premissas=[
                "A PENDENCIA e descontada: o que ja esta comprado e nao chegou "
                "conta como estoque futuro. Sem isso o pedido duplicaria o que "
                "esta a caminho.",
                "A base nao tem campo de categoria terapeutica (molecula e "
                "fabricante vieram vazios). Os agrupamentos oferecidos sao os "
                "que o dado sustenta: curva ABC, faixa de cobertura e marca.",
                f"'{_GRUPO_SEM_LIGACAO}' reune SKUs que existem no arquivo de "
                f"estoque mas nao no sell-out sob o mesmo cadastro — as duas "
                f"fontes trazem EANs diferentes para o mesmo item fisico. Eles "
                f"tem velocidade (vem do proprio arquivo de estoque) mas nao tem "
                f"curva ABC, entao o DDE deles precisa ser decidido a mao.",
                "SKU sem venda no periodo fica fora por padrao: comprar mais do "
                "que nao gira e o oposto do objetivo. Aparece na contagem.",
                "Premissa central: a venda futura segue o ritmo medido. "
                "Sazonalidade, campanha e ruptura de concorrente mudam isso.",
                "Nao considera lote minimo de compra, multiplo de caixa, prazo "
                "de entrega nem validade — nada disso esta na base.",
                "O custo e o de reposicao da foto de estoque; condicao comercial "
                "negociada no pedido pode mudar o valor final.",
            ] + ([
                f"Com valor alvo de R$ {valor_alvo:,.2f}: primeiro passo, cada "
                f"SKU recebe no maximo {teto_por_sku * 100:.0f}% do orcamento "
                f"(protege contra um so item consumir tudo); segundo passo, o "
                f"que sobrar e redistribuido nos mesmos SKUs em ordem de "
                f"urgencia, agora sem teto, ate cada um atingir a necessidade "
                f"real ou o orcamento acabar — o volume que voce definiu nao "
                f"fica parado se houver necessidade real para absorve-lo.",
            ] if valor_alvo is not None else []),
        ).como_dict(),
    }


def exportar_xlsx(dados: dict, nome_cliente: str) -> bytes:
    """Planilha da proposta, pronta para mandar ao comprador.

    Leva as premissas junto numa aba propria: a planilha circula solta por
    e-mail, longe da tela que explica de onde os numeros vieram.
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    if not dados.get("disponivel"):
        raise ValueError(dados.get("motivo", "Sem dados para exportar."))

    wb = Workbook()
    titulo = Font(bold=True, color="FFFFFF")
    fundo = PatternFill("solid", fgColor="7A1420")
    negrito = Font(bold=True)

    ws = wb.active
    ws.title = "Proposta"
    cabecalhos = [
        ("Produto", 46), ("Grupo", 16), ("Filiais", 18), ("Classe estoque", 15),
        ("Estoque atual (un)", 17), ("Pendência (un)", 14),
        ("Venda/mês (un)", 15), ("DDE atual", 11), ("DDE alvo", 10),
        ("Origem do DDE", 14), ("Sugestão (un)", 14),
        ("Custo unit. (R$)", 15), ("Valor (R$)", 15), ("DDE após pedido", 15),
    ]
    for col, (nome, larg) in enumerate(cabecalhos, start=1):
        c = ws.cell(row=1, column=col, value=nome)
        c.font, c.fill = titulo, fundo
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = larg
    ws.freeze_panes = "A2"

    for linha, i in enumerate(dados["itens"], start=2):
        ws.cell(row=linha, column=1, value=i["produto"])
        ws.cell(row=linha, column=2, value=i["grupo"])
        ws.cell(row=linha, column=3, value=", ".join(i["filiais"]))
        ws.cell(row=linha, column=4, value=i["classe_estoque"])
        ws.cell(row=linha, column=5, value=round(i["estoque_atual_un"]))
        ws.cell(row=linha, column=6, value=round(i["pendencia_un"]))
        ws.cell(row=linha, column=7, value=round(i["venda_mes"]))
        c = ws.cell(row=linha, column=8,
                    value=round(i["dde_atual"]) if i["dde_atual"] is not None else None)
        ws.cell(row=linha, column=9, value=round(i["dde_alvo"]))
        ws.cell(row=linha, column=10, value=i["dde_origem"])
        ws.cell(row=linha, column=11, value=round(i["sugestao_un"]))
        c = ws.cell(row=linha, column=12, value=round(i["custo_unitario"], 4))
        c.number_format = "#,##0.0000"
        c = ws.cell(row=linha, column=13, value=round(i["sugestao_valor"], 2))
        c.number_format = "#,##0.00"
        ws.cell(row=linha, column=14,
                value=round(i["dde_apos_pedido"]) if i["dde_apos_pedido"] else None)

    fim = len(dados["itens"]) + 2
    ws.cell(row=fim, column=1, value="TOTAL").font = negrito
    ws.cell(row=fim, column=11, value=round(dados["total_unidades"])).font = negrito
    c = ws.cell(row=fim, column=13, value=round(dados["total_valor"], 2))
    c.font, c.number_format = negrito, "#,##0.00"

    # --- Resumo por grupo ---
    wg = wb.create_sheet("Por grupo")
    for col, nome in enumerate(["Grupo", "DDE alvo", "SKUs", "Unidades", "Valor (R$)"], 1):
        c = wg.cell(row=1, column=col, value=nome)
        c.font, c.fill = titulo, fundo
        wg.column_dimensions[get_column_letter(col)].width = 20
    for linha, g in enumerate(dados["grupos"], start=2):
        wg.cell(row=linha, column=1, value=g["grupo"])
        wg.cell(row=linha, column=2, value=round(g["dde_alvo"]))
        wg.cell(row=linha, column=3, value=g["skus"])
        wg.cell(row=linha, column=4, value=round(g["unidades"]))
        c = wg.cell(row=linha, column=5, value=round(g["valor"], 2))
        c.number_format = "#,##0.00"

    # --- Premissas: a planilha viaja sem a tela que explica os numeros ---
    wp = wb.create_sheet("Como foi calculado")
    wp.column_dimensions["A"].width = 34
    wp.column_dimensions["B"].width = 96
    linha = 1

    def _par(rot: str, val: Any, bold: bool = False) -> None:
        nonlocal linha
        a = wp.cell(row=linha, column=1, value=rot)
        b = wp.cell(row=linha, column=2, value=str(val))
        b.alignment = Alignment(wrap_text=True, vertical="top")
        if bold:
            a.font = b.font = negrito
        linha += 1

    _par("Cliente", nome_cliente, True)
    _par("Foto de estoque", dados["data_ref"])
    _par("Fórmula", dados["calculo"]["formula"])
    linha += 1
    for rot, val in (dados["calculo"].get("valores") or {}).items():
        _par(rot, val)
    linha += 1
    wp.cell(row=linha, column=1, value="PREMISSAS").font = negrito
    linha += 1
    for pr in dados["calculo"].get("premissas") or []:
        _par("", pr)

    if dados.get("corte"):
        linha += 1
        wp.cell(row=linha, column=1, value="ORÇAMENTO").font = negrito
        linha += 1
        c = dados["corte"]
        _par("Valor alvo", f"R$ {c['valor_alvo']:,.2f}")
        _par("Necessidade cheia", f"R$ {c['necessidade_total']:,.2f}")
        _par("Atendido nesta proposta", f"R$ {c['atendido']:,.2f}")
        _par("Necessidade não atendida", f"R$ {c['necessidade_nao_atendida']:,.2f}")
        _par("Teto por SKU", f"{c['teto_por_sku_pct']:.0f}% "
                             f"(R$ {c['teto_por_sku_valor']:,.2f})")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
